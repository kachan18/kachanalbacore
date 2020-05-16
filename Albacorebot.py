import discord
import asyncio
import random
import datetime
import openpyxl
import os

client = discord.Client()


@client.event
async def on_ready():
    print("========================================")
    print("ID : ", client.user.id)
    print("Project Bot. Albacore")
    print("========================================")
    print("ready")
    game = discord.Game("장난을 준비")
    await client.change_presence(status=discord.Status.online, activity=game)


@client.event
async def on_reaction_add(reaction, user):
    channel = reaction.message.channel

    # 통아저씨
    if str(channel) == "통아저씨" and str(reaction.message.author.id) == str(client.user.id) != str(user.id):
        if str(reaction.emoji) == '🗡️':
            if len(reaction.message.embeds) >= 1:
                if reaction.message.embeds[0].title == "원하는 만큼 클릭하세요":
                    rand = random.randint(1, 20)
                    print(rand)
                    if rand == 1:
                        await reaction.message.edit(content=user.mention, allowed_mention="true")
                        embedtong = discord.Embed(title="안타깝게도 알바코어가 힘차게 발사되었습니다.",
                                                  description="당첨자 : %s \n다시 시작하기 : ▶️" % user.name,
                                                  color=0xf15f5f)
                        embedtong.set_image(
                            url="https://images2.imgbox.com/75/14/Z0xw71mT_o.png")
                        await reaction.message.edit(embed=embedtong)
                        await reaction.message.clear_reactions()
                        await reaction.message.add_reaction('▶️')
                    else:
                        embedtong = discord.Embed(title="원하는 만큼 클릭하세요",
                                                  description="칼 꽂기 : :dagger: \n중지버튼 : :stop_button: \n 언제 걸릴지는 랜덤입니다.",
                                                  color=0xf15f5f)
                        embedtong.add_field(name="방금 눌른 사람", value=user.mention, inline=False)
                        embedtong.set_image(
                            url="https://images2.imgbox.com/3c/5c/Y5qAM09P_o.png")
                        await reaction.message.edit(embed=embedtong)
                        await reaction.message.remove_reaction('🗡️', member=user)
        if str(reaction.emoji) == '▶️':
            if len(reaction.message.embeds) >= 1:
                if reaction.message.embeds[0].title == "게임이 중단되어 알바코어가 안도합니다." or reaction.message.embeds[0].title == "안타깝게도 알바코어가 힘차게 발사되었습니다.":
                    await reaction.message.edit(content="@%s" % user.name)
                    embedtong = discord.Embed(title="원하는 만큼 클릭하세요",
                                              description="칼 꽂기 : :dagger: \n중지버튼 : :stop_button: \n 언제 걸릴지는 랜덤입니다.",
                                              color=0xf15f5f)
                    embedtong.add_field(name="시작한 사람", value="@%s" % user.name, inline=False)
                    embedtong.set_image(
                        url="https://images2.imgbox.com/3c/5c/Y5qAM09P_o.png")
                    await reaction.message.edit(embed=embedtong)
                    await reaction.message.clear_reactions()
                    await reaction.message.add_reaction('🗡️')
                    await reaction.message.add_reaction('⏹️')
        if str(reaction.emoji) == '⏹️':
            if len(reaction.message.embeds) >= 1:
                if reaction.message.embeds[0].title == "원하는 만큼 클릭하세요":
                    await reaction.message.edit(content=user.mention, allowed_mention="true")
                    embedtong = discord.Embed(title="게임이 중단되어 알바코어가 안도합니다.",
                                              description="중지자 : %s \n다시 시작하기 : ▶️" % user.name,
                                              color=0xf15f5f)
                    embedtong.set_image(
                        url="https://images2.imgbox.com/e8/cb/DL8EsYZQ_o.png")
                    await reaction.message.edit(embed=embedtong)
                    await reaction.message.clear_reactions()
                    await reaction.message.add_reaction('▶️')

    # 러시안룰렛
    if str(channel) == "러시안룰렛" and str(reaction.message.author.id) == str(client.user.id) != str(user.id):
        file = openpyxl.load_workbook("러시안룰렛.xlsx")
        rr = file.active
        rrn = 0
        for i in range(2, 100):
            if rr["A"+str(i)].value == "-":
                rrn = i - 2
                break
        rrj = []

        if str(reaction.emoji) == '🔫':
            if len(reaction.message.embeds) >= 1:
                if reaction.message.embeds[0].title == "러시안룰렛 대기중":
                    for i in range(2, 100): #100명 이후는 에러나니까 더많으면 숫자 높이기.
                        if rr["A"+str(i)].value == str(user.id):
                            rr["A"+str(i)] = "-"
                            rr["B"+str(i)] = "-"
                            for j in range(i, 100):
                                if rr["A"+str(j)].value == "-":
                                    rr["A" + str(j)] = rr["A" + str(j+1)].value
                                    rr["B" + str(j)] = rr["B" + str(j+1)].value
                                    rr["A" + str(j+1)] = "-"
                                    rr["B" + str(j+1)] = "-"
                            file.save("러시안룰렛.xlsx")
                            break
                        if rr["A"+str(i)].value == "-":
                            rr["A"+str(i)].value = str(user.id)
                            rr["B"+str(i)].value = str(user.name)
                            file.save("러시안룰렛.xlsx")
                            break
                    file = openpyxl.load_workbook("러시안룰렛.xlsx")
                    rr = file.active
                    for i in range(2, 100):
                        if rr["A" + str(i)].value == "-":
                            break
                        rrj.append(rr["B" + str(i)].value)
                    if rrj:
                        rrj = ", ".join(rrj)
                    embedrr = discord.Embed(title="러시안룰렛 대기중",
                                            description="지휘관, 왔어? 좋아, 러시안룰렛을 할거다 이거지? 규칙정도는 알고 있지?",
                                            color=0xf15f5f)
                    embedrr.add_field(name="```규칙```", value="자기 차례가 오면 🔫을 눌러! 운이 없으면? 가는거라구~♪\n그렇게 최후의 1인이 승자가 되는거야!", inline=False)
                    embedrr.add_field(name="```참가 신청```", value="🔫을 누르면 신청이 가능해. 다시 눌러 취소도 할 수 있고.", inline=False)
                    embedrr.add_field(name="```게임 시작```", value="인원 참가가 완료되고 ▶️을 누르면 시작할 수 있어. 적어도 두명은 있어야 된다구?", inline=False)
                    if rrj:
                        embedrr.add_field(name="```참가자```", value="%s" % rrj, inline=False)
                    embedrr.set_thumbnail(
                        url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                    await reaction.message.edit(embed=embedrr)
                    await reaction.message.remove_reaction('🔫', member=user)
                if reaction.message.embeds[0].title == "러시안룰렛 진행중":
                    if rr["A"+str(rr["C2"].value)].value == str(user.id):
                        if rr["C4"].value != rr["C6"].value:
                            if rr["A"+str(rr["C2"].value+1)].value == "-":
                                rr["C2"] = 2
                            else:
                                rr["C2"] = rr["C2"].value + 1
                            rr["C4"] = rr["C4"].value + 1
                            embedrr = discord.Embed(title="러시안룰렛 진행중",
                                                    description="이번에는 운이 좋았네. 과연 누구 차례에... 히히~♪",
                                                    color=0xf15f5f)
                            embedrr.add_field(name="```다음 차례```", value="%s" % rr["B" + str(rr["C2"].value)].value, inline=False)
                            for i in range(2, 100):
                                if rr["A" + str(i)].value == "-":
                                    break
                                rrj.append(rr["B" + str(i)].value)
                            if rrj:
                                rrj = ", ".join(rrj)
                            if rrj:
                                embedrr.add_field(name="```생존자```", value="%s" % rrj, inline=False)
                            embedrr.set_thumbnail(
                                url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                            file.save("러시안룰렛.xlsx")
                            await reaction.message.edit(embed=embedrr)
                            await reaction.message.remove_reaction('🔫', member=user)
                        # 당첨!
                        else:
                            embedrr = discord.Embed(title="러시안룰렛 진행중",
                                                    description="서프라이즈~! 아깝게 됬네 지휘관...히히~♪\n자 그럼, 다시 시작해보자구!",
                                                    color=0xf15f5f)
                            embedrr.add_field(name="```당첨``", value="%s" % rr["B" + str(rr["C2"].value)].value, inline=False)

                            for i in range(2, 50):
                                if rr["A" + str(i)].value == str(user.id):
                                    rr["A" + str(i)].value = "-"
                                    rr["B" + str(i)].value = "-"
                                    for j in range(i, 50):
                                        if rr["A" + str(j)].value == "-":
                                            rr["A" + str(j)].value = rr["A" + str(j + 1)].value
                                            rr["B" + str(j)].value = rr["B" + str(j + 1)].value
                                            rr["A" + str(j + 1)].value = "-"
                                            rr["B" + str(j + 1)].value = "-"
                                    break
                            rrn = rrn - 1
                            if rrn == 1:
                                embedrr = discord.Embed(title="러시안룰렛 종료",
                                                    description="게임이 끝났어! 최후의 1명에게 박수~☆",
                                                    color=0xf15f5f)
                                embedrr.add_field(name="```우승자```", value="%s" % rr["B2"].value, inline=False)
                                embedrr.set_thumbnail(
                                    url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                                udata = openpyxl.load_workbook("Udata.xlsx")
                                ud = udata.active
                                for i in range(2, 50):
                                    if str(ud["A"+str(i)].value) == rr["A2"]:
                                        ud["E"+str(i)] = ud["E"+str(i)].value + 50
                                        break
                                    if str(ud["A" + str(i)].value) == "-":
                                        ud["A" + str(i)] = rr["A2"].value
                                        ud["E" + str(i)] = 1050
                                        break
                                rr["A2"] = "-"
                                rr["B2"] = "-"
                                file.save("러시안룰렛.xlsx")
                                udata.save("Udata.xlsx")
                                await reaction.message.edit(embed=embedrr)
                                await reaction.message.clear_reactions()
                            else:
                                rr["C4"] = 1
                                rr["C6"] = random.randint(3, 2 * rrn)
                                rr["C2"] = random.randint(2, 1 + rrn)
                                embedrr.add_field(name="```다음 차례```", value="%s" % rr["B" + str(rr["C2"].value)].value, inline=False)
                                for i in range(2, 100):
                                    if rr["A" + str(i)].value == "-":
                                        break
                                    rrj.append(rr["B" + str(i)].value)
                                if rrj:
                                    rrj = ", ".join(rrj)
                                if rrj:
                                    embedrr.add_field(name="```생존자```", value="%s" % rrj, inline=False)
                                embedrr.set_thumbnail(
                                    url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                                file.save("러시안룰렛.xlsx")
                                await reaction.message.edit(embed=embedrr)
                                await reaction.message.remove_reaction('🔫', member=user)
                    # 차례 오류시
                    else:
                        embedrr = discord.Embed(title="러시안룰렛 진행중",
                                                description="과연 누가 이길까나~ 히히~♪",
                                                color=0xf15f5f)
                        embedrr.add_field(name="```다음 차례```", value="%s" % rr["B"+str(rr["C2"].value)].value, inline=False)
                        for i in range(2, 100):
                            if rr["A" + str(i)].value == "-":
                                break
                            rrj.append(rr["B" + str(i)].value)
                        if rrj:
                            rrj = ", ".join(rrj)
                        if rrj:
                            embedrr.add_field(name="```생존자```", value="%s" % rrj, inline=False)
                        embedrr.add_field(name="```오류```", value="[ %s ] 지휘관 차례가 아니잖아! 순서 정도는 지켜달라구!" % user.name, inline=False)
                        embedrr.set_thumbnail(
                            url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                        await reaction.message.edit(embed=embedrr)
                        await reaction.message.remove_reaction('🔫', member=user)
        if str(reaction.emoji) == '▶️':
            if len(reaction.message.embeds) >= 1:
                if reaction.message.embeds[0].title == "러시안룰렛 대기중":
                    if rr["A"+str(3)].value == "-":
                        embedrr = discord.Embed(title="러시안룰렛 대기중",
                                                description="지휘관, 왔어? 좋아, 러시안룰렛을 할거다 이거지? 규칙정도는 알고 있지?",
                                                color=0xf15f5f)
                        embedrr.add_field(name="```규칙```", value="자기 차례가 오면 🔫을 눌러! 운이 없으면? 가는거라구~♪\n그렇게 최후의 1인이 승자가 되는거야!", inline=False)
                        embedrr.add_field(name="```참가 신청```", value="🔫을 누르면 신청이 가능해. 다시 눌러 취소도 할 수 있고.", inline=False)
                        embedrr.add_field(name="```게임 시작```", value="인원 참가가 완료되고 ▶️을 누르면 시작할 수 있어. 적어도 두명은 있어야 된다구?", inline=False)
                        for i in range(2, 100):
                            if rr["A" + str(i)].value == "-":
                                break
                            rrj.append(rr["B" + str(i)].value)
                        if rrj:
                            rrj = ", ".join(rrj)
                        if rrj:
                            embedrr.add_field(name="```참가자```", value="%s" % rrj, inline=False)
                        embedrr.set_thumbnail(
                            url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                        embedrr.add_field(name="```오류```", value="위에도 말했지만, 적어도 두명은 있어야 된다구?", inline=False)
                        await reaction.message.edit(embed=embedrr)
                        await reaction.message.remove_reaction('▶️', member=user)
                    else:
                        embedrr = discord.Embed(title="러시안룰렛 진행중",
                                                description="좋아, 이제 시작해 볼까? 행운을 빌어.",
                                                color=0xf15f5f)
                        rrp = random.randint(2, 1+rrn)
                        rrd = random.randint(3, 2 * rrn)
                        embedrr.add_field(name="```처음 차례```", value="%s" % rr["B"+str(rrp)].value, inline=False)
                        for i in range(2, 100):
                            if rr["A" + str(i)].value == "-":
                                break
                            rrj.append(rr["B" + str(i)].value)
                        if rrj:
                            rrj = ", ".join(rrj)
                        if rrj:
                            embedrr.add_field(name="```생존자```", value="%s" % rrj, inline=False)
                        rr["C2"] = rrp
                        rr["C4"] = 1
                        rr["C6"] = rrd
                        file.save("러시안룰렛.xlsx")

                        embedrr.set_thumbnail(
                            url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                        await reaction.message.edit(embed=embedrr)
                        await reaction.message.clear_reactions()
                        await reaction.message.add_reaction('🔫')
                        await reaction.message.add_reaction('⏹️')
        if str(reaction.emoji) == '⏹️':
            if len(reaction.message.embeds) >= 1:
                if reaction.message.embeds[0].title == "러시안룰렛 진행중":
                    await reaction.message.edit(content=user.mention, allowed_mention="true")
                    embedrr = discord.Embed(title="러시안룰렛 강제종료",
                                              description="중지자 : %s" % user.name,
                                              color=0xf15f5f)
                    embedrr.set_thumbnail(
                        url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                    rr["C2"] = "-"
                    rr["C4"] = "-"
                    rr["C6"] = "-"
                    for i in range(2, 1+rrn):
                        rr["A"+str(i)] = "-"
                        rr["B"+str(i)] = "-"
                    file.save("러시안룰렛.xlsx")
                    await reaction.message.edit(embed=embedrr)
                    await reaction.message.clear_reactions()


@client.event
async def on_message(message):

    id = message.author.id
    channel = message.channel

    if message.author.bot:  # 봇은 기본적으로 무시하지만?
        channel = message.channel
        if str(channel) == "러시안룰렛" and str(message.author.id) == str(client.user.id):
            if len(message.embeds) >= 1:
                if message.embeds[0].description == "지휘관, 왔어? 좋아, 러시안룰렛을 할거다 이거지? 규칙정도는 알고 있는거지? 모른다고? 어쩔 수 없네~":
                    await message.add_reaction('🔫')
                    await message.add_reaction('▶️')
        if str(channel) == "통아저씨" and str(message.author.id) == str(client.user.id):
            if len(message.embeds) >= 1:
                if message.embeds[0].title == "원하는 만큼 클릭하세요":
                    await message.add_reaction('🗡️')
                    await message.add_reaction('⏹️')
        return None

    if message.content.startswith("!알바코어"):
        cmdline = message.content.split(' ')

        if len(cmdline) == 1:
            await channel.send("하항? 혹시 지휘관... 작은 게 취향인 거야?")
        elif len(cmdline) >= 2:
            if cmdline[1] == "도움말":
                embedhelp = discord.Embed(title="지휘관, 궁금한 거라도 있는거야?", description="부탁 한 번 들어주면 알려줄게~",
                                          color=0xf15f5f)
                embedhelp.set_thumbnail(
                    url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                embedhelp.add_field(name="!알바코어", value="오늘도 신나게 장난...이 아니지, 열심히 일할게~!", inline=False)
                embedhelp.add_field(name="!알바코어 통아저씨", value="저기... 왜 나를 통안에 넣은거야...? 나가도 돼?", inline=False)
                embedhelp.add_field(name="!알바코어 러시안룰렛", value="지휘관, 러시안룰렛 해볼래?", inline=False)
                embedhelp.add_field(name="!알바코어 출석체크", value="출석체크!", inline=False)
                embedhelp.add_field(name="!알바코어 크레딧", value="지휘관, 크레딧에 대해 궁금하다고?", inline=False)
                await channel.send(embed=embedhelp)

            elif cmdline[1] == "통아저씨":
                if str(channel) == "통아저씨":
                    embedtong = discord.Embed(title="원하는 만큼 클릭하세요",
                                              description="칼 꽂기 : :dagger: \n중지버튼 : :stop_button: \n 언제 걸릴지는 랜덤입니다.",
                                              color=0xf15f5f)
                    embedtong.set_image(
                        url="https://images2.imgbox.com/3c/5c/Y5qAM09P_o.png")
                    await channel.send(embed=embedtong)
                else:
                    await channel.send("이 기능은 #통아저씨 채널에서만 사용이 가능하다구")

            elif cmdline[1] == "러시안룰렛":
                if str(channel) == "러시안룰렛":
                    embedrr = discord.Embed(title="러시안룰렛 대기중",
                                              description="지휘관, 왔어? 좋아, 러시안룰렛을 할거다 이거지? 규칙정도는 알고 있는거지? 모른다고? 어쩔 수 없네~",
                                              color=0xf15f5f)
                    embedrr.add_field(name="```규칙```", value="자기 차례가 오면 🔫을 눌러! 운이 없으면? 가는거라구~♪\n그렇게 최후의 1인이 승자가 되는거야!\n과연 몇 번째 약실에 총알이 들어있을까?", inline=False)
                    embedrr.add_field(name="```참가 신청```", value="🔫을 누르면 신청이 가능해. 다시 눌러 취소도 할 수 있고.", inline=False)
                    embedrr.add_field(name="```게임 시작```", value="인원 참가가 완료되고 ▶️을 누르면 시작할 수 있어. 적어도 두명은 있어야 된다구?", inline=False)
                    embedrr.set_thumbnail(
                        url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                    await channel.send(embed=embedrr)
                else:
                    await channel.send("이 기능은 #러시안룰렛 채널에서만 사용이 가능하다구")

            elif cmdline[1] == "정보":
                embed = discord.Embed(title="알바코어봇 v0.2",
                                      description="게임과 장난을 위한 봇\n제작자 : Admiral. 레이나",
                                      color=0xf15f5f)
                embed.set_image(
                    url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                await channel.send(embed=embed)

            elif cmdline[1] == "출석체크":
                file = openpyxl.load_workbook("Udata.xlsx")
                sheet = file.active
                for i in range(2, 51):
                    if str(message.author.id) == sheet["A" + str(i)].value:
                        if int(sheet["B" + str(i)].value) < int(datetime.datetime.now().strftime("%Y%m%d")):
                            cooldown = datetime.datetime.now()
                            sheet["B" + str(i)].value = cooldown.strftime("%Y%m%d")
                            if sheet["B" + str(i)].value == int(datetime.datetime.now().strftime("%Y%m%d")) - 1:
                                sheet["C" + str(i)] = sheet["C" + str(i)].value + 1
                            else:
                                sheet["C" + str(i)] = 1
                            sheet["D" + str(i)] = sheet["D" + str(i)].value + 1
                            sheet["E" + str(i)] = sheet["E" + str(i)].value + 100
                            file.save("Udata.xlsx")
                            embed = discord.Embed(title="출석체크", description="지휘관, 오늘도 신나게 장난...이 아니지, 열심히 보내보자구~!\n출석체크 보상으로 100 크레딧도 줄게.",
                                                      color=0xf15f5f)
                            embed.set_thumbnail(
                                url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                            embed.add_field(name="```출석자```", value="[ %s ] 지휘관" % message.author.mention, inline=False)
                            embed.add_field(name="```출석일자```", value="%s" % cooldown.strftime("%Y / %m / %d"), inline=False)
                            embed.add_field(name="```연속 출석일수```", value="%d 일" % sheet["D" + str(i)].value, inline=False)
                            embed.add_field(name="```총 출석일수```", value="%d 일" % sheet["D" + str(i)].value, inline=False)
                            await channel.send(embed=embed)
                        else:
                            await channel.send("```[%s] 지휘관, 오늘은 벌써 출석체크를 했다구```" % message.author)
                            break
                    if sheet["A" + str(i)].value == "-":
                        sheet["A" + str(i)] = str(message.author.id)
                        cooldown = datetime.datetime.now()
                        sheet["B" + str(i)] = cooldown.strftime("%Y%m%d")
                        sheet["C" + str(i)] = 1
                        sheet["D" + str(i)] = 1
                        sheet["E" + str(i)] = 1100
                        file.save("Udata.xlsx")
                        embed = discord.Embed(title="출석체크", description="지휘관, 오늘도 신나게 장난...이 아니지, 열심히 보내보자구~!\n출석체크 보상으로 100 크레딧도 줄게.",
                                              color=0xf15f5f)
                        embed.set_thumbnail(
                            url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                        embed.add_field(name="```출석자```", value="[ %s ] 지휘관" % message.author.mention, inline=False)
                        embed.add_field(name="```출석일자```", value="%s" % cooldown.strftime("%Y / %m / %d"), inline=False)
                        embed.add_field(name="```연속 출석일수```", value="%d 일" % sheet["D" + str(i)].value, inline=False)
                        embed.add_field(name="```총 출석일수```", value="%d 일" % sheet["D" + str(i)].value, inline=False)
                        await channel.send(embed=embed)
                        break

            elif cmdline[1] == "크레딧":
                if len(cmdline) == 2:
                    embed = discord.Embed(title="크레딧", description="지휘관, 크레딧에 대해 궁금하다고?",
                                          color=0xf15f5f)
                    embed.set_thumbnail(
                        url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                    embed.add_field(name="```크레딧 확인```", value="지휘관이 얼마나 가지고 있는지 알 수 있어.", inline=False)
                    await channel.send(embed=embed)
                else:
                    if cmdline[2] == "확인":
                        file = openpyxl.load_workbook("Udata.xlsx")
                        sheet = file.active
                        embed = discord.Embed(title="크레딧", description="지휘관, 아이스크림 사주면, 누군가의 '남들에게 걸려선 안 되는 일'은 잊어버릴지도~",
                                              color=0xf15f5f)
                        embed.set_thumbnail(
                            url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                        embed.add_field(name="```보유자```", value="[ %s ] 지휘관" % message.author.mention, inline=False)
                        for i in range(2, 51):
                            if str(message.author.id) == sheet["A" + str(i)].value:
                                embed.add_field(name="```보유량```", value="%s 크레딧" % sheet["E" + str(i)].value, inline=False)
                                break
                            if sheet["A" + str(i)].value == "-":
                                sheet["A" + str(i)] = str(message.author.id)
                                embed.add_field(name="```보유량```", value="%s 크레딧" % sheet["E" + str(i)].value, inline=False)
                                break
                        await channel.send(embed=embed)
                    elif cmdline[2] == "디버그용추가":
                        file = openpyxl.load_workbook("Udata.xlsx")
                        sheet = file.active
                        for i in range(2, 51):
                            if str(message.author.id) == sheet["A" + str(i)].value:
                                break
                        if sheet["F" + str(i)].value == "관리자":
                            if len(cmdline) == 4:
                                embed = discord.Embed(title="크레딧", description="디버그용 추가 기능 -로 제거도 가능",
                                                      color=0xf15f5f)
                                embed.set_thumbnail(
                                    url="https://images2.imgbox.com/8d/01/GdvzdwSj_o.png")
                                embed.add_field(name="```보유자```", value="[ %s ] 지휘관" % message.author.mention, inline=False)
                                for i in range(2, 51):
                                    if str(message.author.id) == sheet["A" + str(i)].value:
                                        embed.add_field(name="```이전 보유량```", value="%s 크레딧" % sheet["E" + str(i)].value, inline=False)
                                        break
                                sheet["E" + str(i)] = sheet["E" + str(i)].value + int(cmdline[3])
                                file.save("Udata.xlsx")
                                embed.add_field(name="```현재 보유량```", value="%s 크레딧" % sheet["E" + str(i)].value, inline=False)
                                await channel.send(embed=embed)
                            else:
                                await channel.send("지휘관, 필요한 정보 중 일부가 빠졌는데?")
                        else:
                            await channel.send("지휘관, 필요한 권한이 없는데?")
                    else:
                        await channel.send("그런 기능 들어본적 없는데?")
            else:
                await channel.send("뭘 원하는거야 지휘관?")
    if message.content.startswith("!알바코난"):
        await channel.send("누가 알바코난이야! 알바코어라고!")

access_token = os.environ["BOT_TOKEN"]
client.run(access_token)